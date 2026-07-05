import logging
import re
from urllib.parse import DefragResult
from django.shortcuts import render, HttpResponse, redirect, get_object_or_404
from django.http import FileResponse, JsonResponse, Http404, HttpResponse
from django.core import serializers
from itertools import chain
from .forms import SearchForm, ModelSearchForm
import json
import folium
from folium.plugins import HeatMap
import math
import geocoder
from folium import plugins
from datetime import datetime

from agmp_app.models import *
from django.db.models import Avg, Min, Max, Count, Q, F
import pandas as pd
from collections import Counter, defaultdict
from django_pandas.io import read_frame
from django.views.generic.detail import DetailView
from django.views.generic import ListView, TemplateView
import numpy as np
import os

from django.db.models import Subquery, OuterRef
from django.conf import settings
import geopandas as gpd
from shapely.geometry import Point
from fuzzywuzzy import fuzz
from fuzzywuzzy import process

from django.views.decorators.http import require_http_methods

import csv
import io

# ============================================================ # ============================================================


logger = logging.getLogger(__name__)

# Maximum identifiers per batch query submission.
# Reads from settings.AGMP_BATCH_LIMIT (set via env var in settings.py).
BATCH_QUERY_LIMIT = getattr(settings, 'AGMP_BATCH_LIMIT', 50)


HEADER_MAP = {
    '_input': 'Search Input', '_status': 'Status',
    'rs_id': 'RS ID', 'gene_name': 'Gene Name', 'chromosome': 'Chromosome',
    'variant_type': 'Variant Type', 'allele': 'Allele', 'source_db': 'Source DB',
    'gene_id': 'Gene ID', 'gene_function': 'Gene Function', 'function': 'Gene Function',
    'uniprot_ac': 'UniProt AC', 'drug_name': 'Drug Name', 'drug_bank_id': 'DrugBank ID',
    'drug_id': 'Drug ID', 'state': 'State', 'indication': 'Indication',
    'iupac_name': 'IUPAC Name', 'phenotype_name': 'Phenotype Name',
    'drug_assoc_count': 'Drug Assoc. Count', 'phenotype_assoc_count': 'Phenotype Assoc. Count',
    'study_count': 'Study Count', 'variant_count': 'Variant Count', 'gene_count': 'Gene Count',
    'variants_detail': 'Associated Variants', 'drugs_detail': 'Associated Drugs',
    'phenotypes_detail': 'Associated Phenotypes', 'studies_detail': 'Associated Studies',
    'genes_detail': 'Associated Genes', 'countries_detail': 'Study Countries',
}

DETAIL_FIELD_KEYS = {
    'studies_detail': ['title', 'pubmed_id', 'year', 'study_type'],
    'drugs_detail': ['name', 'drug_bank_id', 'state', 'indication'],
    'genes_detail': ['name', 'id', 'chromosome', 'function'],
    'variants_detail': ['id', 'type', 'allele', 'gene'],
    'phenotypes_detail': ['name'],
    'countries_detail': ['country', 'lat', 'lng'],
}


def batch_query_view(request):
    return render(request, 'batch_query.html')


def batch_query_execute(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'})
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'})

    query_type = data.get('query_type', '')
    queries = data.get('queries', [])
    output_fields = data.get('output_fields', [])

    if not query_type:
        return JsonResponse({'success': False, 'error': 'Select a query type'})
    if not queries:
        return JsonResponse({'success': False, 'error': 'Provide at least one search term'})
    if not output_fields:
        return JsonResponse({'success': False, 'error': 'Select at least one output field'})

    seen = set()
    unique_queries = []
    for q in queries:
        q = q.strip()
        if q and q.lower() not in seen:
            seen.add(q.lower())
            unique_queries.append(q)
    total_submitted = len(unique_queries)
    truncated = total_submitted > BATCH_QUERY_LIMIT
    queries = unique_queries[:BATCH_QUERY_LIMIT]

    results, not_found = [], []
    query_func = {
        'variant': query_variant_data,
        'gene': query_gene_data,
        'drug': query_drug_data,
        'phenotype': query_phenotype_data,
    }.get(query_type)

    for query in queries:
        row = query_func(query, output_fields) if query_func else None
        if row:
            row['_input'] = query
            row['_status'] = 'found'
            results.append(row)
        else:
            results.append({'_input': query, '_status': 'not_found'})
            not_found.append(query)

    return JsonResponse({
        'success': True, 'results': results,
        'total': len(results), 'found': len(results) - len(not_found),
        'not_found_count': len(not_found), 'not_found_list': not_found,
        'limit': BATCH_QUERY_LIMIT, 'truncated': truncated,
        'total_submitted': total_submitted,
    })


# ── Query helpers ───────────────────────────

def query_variant_data(rs_id, fields):
    try:
        variant = Variantagmp.objects.filter(
            Q(rs_id__iexact=rs_id) | Q(rs_id__icontains=rs_id)
        ).select_related('geneagmp', 'drugagmp', 'phenotypeagmp').first()
        if not variant:
            return None
        result = {}
        if 'rs_id' in fields:
            result['rs_id'] = variant.rs_id or ''
        if 'variant_type' in fields:
            result['variant_type'] = variant.variant_type or ''
        if 'allele' in fields:
            result['allele'] = variant.allele or ''
        if 'source_db' in fields:
            result['source_db'] = variant.source_db or ''
        if 'gene_name' in fields:
            result['gene_name'] = variant.geneagmp.gene_name if variant.geneagmp else ''
        if 'chromosome' in fields:
            result['chromosome'] = variant.geneagmp.chromosome if variant.geneagmp else ''
        if 'gene_id' in fields:
            result['gene_id'] = variant.geneagmp.gene_id if variant.geneagmp else ''
        if 'gene_function' in fields:
            result['gene_function'] = variant.geneagmp.function if variant.geneagmp else ''
        if 'uniprot_ac' in fields:
            result['uniprot_ac'] = variant.geneagmp.uniprot_ac if variant.geneagmp else ''

        all_variants = Variantagmp.objects.filter(rs_id__iexact=rs_id).select_related('drugagmp', 'phenotypeagmp')

        if 'drugs_detail' in fields:
            drugs, seen = [], set()
            for v in all_variants:
                if v.drugagmp and v.drugagmp.drug_bank_id not in seen:
                    seen.add(v.drugagmp.drug_bank_id)
                    drugs.append({'name': v.drugagmp.drug_name or '', 'drug_bank_id': v.drugagmp.drug_bank_id or '', 'indication': (v.drugagmp.indication or '')[:100]})
            result['drugs_detail'] = drugs[:30]

        if 'phenotypes_detail' in fields:
            phenotypes, seen = [], set()
            for v in all_variants:
                if v.phenotypeagmp and v.phenotypeagmp.name not in seen:
                    seen.add(v.phenotypeagmp.name)
                    phenotypes.append({'name': v.phenotypeagmp.name})
            result['phenotypes_detail'] = phenotypes[:50]

        if 'studies_detail' in fields or 'countries_detail' in fields:
            variant_studies = VariantStudyagmp.objects.filter(variantagmp__rs_id__iexact=rs_id).select_related('studyagmp')
            if 'studies_detail' in fields:
                studies, seen = [], set()
                for vs in variant_studies:
                    if vs.studyagmp and vs.studyagmp.publication_id not in seen:
                        seen.add(vs.studyagmp.publication_id)
                        studies.append({'title': (vs.studyagmp.title or '')[:150], 'pubmed_id': vs.studyagmp.publication_id or '', 'year': vs.studyagmp.publication_year or '', 'study_type': vs.studyagmp.study_type or ''})
                result['studies_detail'] = studies[:20]
            if 'countries_detail' in fields:
                countries = []
                for vs in variant_studies[:10]:
                    countries.extend(extract_countries(vs))
                result['countries_detail'] = countries[:30]

        if 'drug_assoc_count' in fields:
            result['drug_assoc_count'] = all_variants.exclude(source_db__in=["DisGeNET", "GWAS Catalog"]).exclude(drugagmp__isnull=True).count()
        if 'phenotype_assoc_count' in fields:
            result['phenotype_assoc_count'] = all_variants.filter(source_db__in=["DisGeNET", "GWAS Catalog"]).exclude(phenotypeagmp__isnull=True).count()
        if 'study_count' in fields:
            result['study_count'] = VariantStudyagmp.objects.filter(variantagmp__rs_id__iexact=rs_id).values('studyagmp__publication_id').distinct().count()
        return result
    except Exception as e:
        logger.error(f"query_variant_data error: {e}")
        return None


def query_gene_data(gene_input, fields):
    try:
        gene = Geneagmp.objects.filter(Q(gene_id__iexact=gene_input) | Q(gene_name__iexact=gene_input)).first()
        if not gene:
            return None
        result = {}
        if 'gene_id' in fields:
            result['gene_id'] = gene.gene_id or ''
        if 'gene_name' in fields:
            result['gene_name'] = gene.gene_name or ''
        if 'chromosome' in fields:
            result['chromosome'] = gene.chromosome or ''
        if 'function' in fields:
            result['function'] = gene.function or ''
        if 'uniprot_ac' in fields:
            result['uniprot_ac'] = gene.uniprot_ac or ''

        variants = Variantagmp.objects.filter(geneagmp=gene).select_related('drugagmp', 'phenotypeagmp')

        if 'variants_detail' in fields:
            vl, seen = [], set()
            for v in variants:
                if v.rs_id and v.rs_id not in seen:
                    seen.add(v.rs_id)
                    vl.append({'id': v.rs_id, 'type': v.variant_type or '', 'allele': v.allele or ''})
            result['variants_detail'] = vl[:50]
        if 'drugs_detail' in fields:
            dl, seen = [], set()
            for v in variants:
                if v.drugagmp and v.drugagmp.drug_bank_id not in seen:
                    seen.add(v.drugagmp.drug_bank_id)
                    dl.append({'name': v.drugagmp.drug_name or '', 'drug_bank_id': v.drugagmp.drug_bank_id or '', 'state': v.drugagmp.state or ''})
            result['drugs_detail'] = dl[:30]
        if 'phenotypes_detail' in fields:
            pl, seen = [], set()
            for v in variants:
                if v.phenotypeagmp and v.phenotypeagmp.name not in seen:
                    seen.add(v.phenotypeagmp.name)
                    pl.append({'name': v.phenotypeagmp.name})
            result['phenotypes_detail'] = pl[:50]

        if 'studies_detail' in fields or 'countries_detail' in fields:
            vs_qs = VariantStudyagmp.objects.filter(variantagmp__geneagmp=gene).select_related('studyagmp')
            if 'studies_detail' in fields:
                sl, seen = [], set()
                for vs in vs_qs:
                    if vs.studyagmp and vs.studyagmp.publication_id not in seen:
                        seen.add(vs.studyagmp.publication_id)
                        sl.append({'title': (vs.studyagmp.title or '')[:150], 'pubmed_id': vs.studyagmp.publication_id or '', 'year': vs.studyagmp.publication_year or '', 'study_type': vs.studyagmp.study_type or ''})
                result['studies_detail'] = sl[:20]
            if 'countries_detail' in fields:
                cl = []
                for vs in vs_qs[:15]:
                    cl.extend(extract_countries(vs))
                seen, unique = set(), []
                for c in cl:
                    if c['country'] not in seen:
                        seen.add(c['country'])
                        unique.append(c)
                result['countries_detail'] = unique[:30]

        if 'variant_count' in fields:
            result['variant_count'] = variants.values('rs_id').distinct().count()
        if 'drug_assoc_count' in fields:
            result['drug_assoc_count'] = variants.exclude(drugagmp__isnull=True).count()
        if 'phenotype_assoc_count' in fields:
            result['phenotype_assoc_count'] = variants.exclude(phenotypeagmp__isnull=True).count()
        if 'study_count' in fields:
            result['study_count'] = VariantStudyagmp.objects.filter(variantagmp__geneagmp=gene).values('studyagmp__publication_id').distinct().count()
        return result
    except Exception as e:
        logger.error(f"query_gene_data error: {e}")
        return None


def query_drug_data(drug_input, fields):
    try:
        drug = Drugagmp.objects.filter(Q(drug_name__iexact=drug_input) | Q(drug_bank_id__iexact=drug_input)).first()
        if not drug:
            return None
        result = {}
        if 'drug_name' in fields:
            result['drug_name'] = drug.drug_name or ''
        if 'drug_bank_id' in fields:
            result['drug_bank_id'] = drug.drug_bank_id or ''
        if 'drug_id' in fields:
            result['drug_id'] = drug.drug_id or ''
        if 'state' in fields:
            result['state'] = drug.state or ''
        if 'indication' in fields:
            result['indication'] = drug.indication or ''
        if 'iupac_name' in fields:
            result['iupac_name'] = drug.iupac_name_seq or ''

        variants = Variantagmp.objects.filter(drugagmp=drug).select_related('geneagmp', 'phenotypeagmp')

        if 'genes_detail' in fields:
            gl, seen = [], set()
            for v in variants:
                if v.geneagmp and v.geneagmp.gene_id not in seen:
                    seen.add(v.geneagmp.gene_id)
                    gl.append({'id': v.geneagmp.gene_id or '', 'name': v.geneagmp.gene_name or '', 'chromosome': v.geneagmp.chromosome or '', 'function': (v.geneagmp.function or '')[:100]})
            result['genes_detail'] = gl[:30]
        if 'variants_detail' in fields:
            vl, seen = [], set()
            for v in variants:
                if v.rs_id and v.rs_id not in seen:
                    seen.add(v.rs_id)
                    vl.append({'id': v.rs_id, 'type': v.variant_type or '', 'gene': v.geneagmp.gene_name if v.geneagmp else ''})
            result['variants_detail'] = vl[:50]
        if 'phenotypes_detail' in fields:
            pl, seen = [], set()
            for v in variants:
                if v.phenotypeagmp and v.phenotypeagmp.name not in seen:
                    seen.add(v.phenotypeagmp.name)
                    pl.append({'name': v.phenotypeagmp.name})
            result['phenotypes_detail'] = pl[:50]
        if 'studies_detail' in fields:
            vs_qs = VariantStudyagmp.objects.filter(variantagmp__drugagmp=drug).select_related('studyagmp')
            sl, seen = [], set()
            for vs in vs_qs:
                if vs.studyagmp and vs.studyagmp.publication_id not in seen:
                    seen.add(vs.studyagmp.publication_id)
                    sl.append({'title': (vs.studyagmp.title or '')[:150], 'pubmed_id': vs.studyagmp.publication_id or '', 'year': vs.studyagmp.publication_year or '', 'study_type': vs.studyagmp.study_type or ''})
            result['studies_detail'] = sl[:20]

        if 'variant_count' in fields:
            result['variant_count'] = variants.values('rs_id').distinct().count()
        if 'gene_count' in fields:
            result['gene_count'] = variants.values('geneagmp__gene_id').distinct().count()
        if 'study_count' in fields:
            result['study_count'] = VariantStudyagmp.objects.filter(variantagmp__drugagmp=drug).values('studyagmp__publication_id').distinct().count()
        return result
    except Exception as e:
        logger.error(f"query_drug_data error: {e}")
        return None


def query_phenotype_data(phenotype_name, fields):
    try:
        variants = Variantagmp.objects.filter(phenotypeagmp__name__iexact=phenotype_name).select_related('geneagmp', 'drugagmp', 'phenotypeagmp')
        if not variants.exists():
            return None
        result = {}
        if 'phenotype_name' in fields:
            result['phenotype_name'] = phenotype_name
        if 'genes_detail' in fields:
            gl, seen = [], set()
            for v in variants:
                if v.geneagmp and v.geneagmp.gene_id not in seen:
                    seen.add(v.geneagmp.gene_id)
                    gl.append({'id': v.geneagmp.gene_id or '', 'name': v.geneagmp.gene_name or '', 'chromosome': v.geneagmp.chromosome or ''})
            result['genes_detail'] = gl[:30]
        if 'variants_detail' in fields:
            vl, seen = [], set()
            for v in variants:
                if v.rs_id and v.rs_id not in seen:
                    seen.add(v.rs_id)
                    vl.append({'id': v.rs_id, 'gene': v.geneagmp.gene_name if v.geneagmp else ''})
            result['variants_detail'] = vl[:50]
        if 'drugs_detail' in fields:
            dl, seen = [], set()
            for v in variants:
                if v.drugagmp and v.drugagmp.drug_bank_id not in seen:
                    seen.add(v.drugagmp.drug_bank_id)
                    dl.append({'name': v.drugagmp.drug_name or '', 'drug_bank_id': v.drugagmp.drug_bank_id or ''})
            result['drugs_detail'] = dl[:30]
        if 'studies_detail' in fields:
            vs_qs = VariantStudyagmp.objects.filter(variantagmp__phenotypeagmp__name__iexact=phenotype_name).select_related('studyagmp')
            sl, seen = [], set()
            for vs in vs_qs:
                if vs.studyagmp and vs.studyagmp.publication_id not in seen:
                    seen.add(vs.studyagmp.publication_id)
                    sl.append({'title': (vs.studyagmp.title or '')[:150], 'pubmed_id': vs.studyagmp.publication_id or '', 'year': vs.studyagmp.publication_year or '', 'study_type': vs.studyagmp.study_type or ''})
            result['studies_detail'] = sl[:20]
        if 'variant_count' in fields:
            result['variant_count'] = variants.values('rs_id').distinct().count()
        if 'gene_count' in fields:
            result['gene_count'] = variants.values('geneagmp__gene_id').distinct().count()
        if 'study_count' in fields:
            result['study_count'] = VariantStudyagmp.objects.filter(variantagmp__phenotypeagmp__name__iexact=phenotype_name).values('studyagmp__publication_id').distinct().count()
        return result
    except Exception as e:
        logger.error(f"query_phenotype_data error: {e}")
        return None


def extract_countries(variant_study):
    countries = []
    if variant_study.country_participant:
        countries.append({'country': variant_study.country_participant, 'lat': variant_study.latitude, 'lng': variant_study.longitude})
    for i in range(1, 31):
        suffix = f'_{i:02d}' if i < 10 else (f'_{i:03d}' if i <= 19 else f'_{i}')
        country = getattr(variant_study, f'country_participant{suffix}', None)
        if country:
            countries.append({'country': country, 'lat': getattr(variant_study, f'latitude{suffix}', None), 'lng': getattr(variant_study, f'longitude{suffix}', None)})
    return countries


# ── Flat CSV export (backward compat) ───────

def _flatten_detail_list(field_type, items):
    if not items:
        return ''
    keys = DETAIL_FIELD_KEYS.get(field_type)
    parts = []
    for item in items:
        if isinstance(item, dict):
            vals = [str(item.get(k, '')) for k in (keys or item.keys()) if item.get(k) and str(item.get(k, '')).lower() != 'nan']
            if vals:
                parts.append(' | '.join(vals))
        else:
            s = str(item)
            if s.lower() != 'nan':
                parts.append(s)
    return '; '.join(parts)


def batch_query_export(request):
    if request.method != 'POST':
        return HttpResponse("POST required", status=405)
    try:
        data = json.loads(request.body)
        results = data.get('results', [])
        columns = data.get('columns', [])
        query_type = data.get('query_type', 'query')
        chunk_index = data.get('chunk_index', 1)
        chunk_total = data.get('chunk_total', 1)
    except json.JSONDecodeError:
        return HttpResponse("Invalid data", status=400)
    if not results:
        return HttpResponse("No results", status=400)
    if not columns:
        columns = list(results[0].keys())

    filename = f"batchquery_{chunk_index}_of_{chunk_total}_{query_type}.csv"
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.write('\ufeff')
    writer = csv.writer(response)
    writer.writerow([HEADER_MAP.get(c, c.replace('_', ' ').title()) for c in columns])
    for row in results:
        csv_row = []
        for col in columns:
            val = row.get(col, '')
            if isinstance(val, list):
                val = _flatten_detail_list(col, val)
            elif val is None:
                val = ''
            csv_row.append(val)
        writer.writerow(csv_row)
    return response


# ── Multi-sheet Excel export ────────────────

def batch_query_export_xlsx(request):
    """Build a multi-sheet .xlsx with normalized tables linked by Search Input."""
    if request.method != 'POST':
        return HttpResponse("POST required", status=405)
    try:
        data = json.loads(request.body)
        results = data.get('results', [])
        main_fields = data.get('main_fields', [])
        detail_fields = data.get('detail_fields', [])
        query_type = data.get('query_type', 'query')
    except json.JSONDecodeError:
        return HttpResponse("Invalid JSON", status=400)
    if not results:
        return HttpResponse("No results to export", status=400)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # Styles
    hdr_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    hdr_fill = PatternFill('solid', fgColor='2E7D32')
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    link_font = Font(name='Arial', color='1565C0', underline='single', size=10)
    found_font = Font(name='Arial', color='2E7D32', bold=True, size=10)
    notfound_font = Font(name='Arial', color='C62828', bold=True, size=10)
    data_font = Font(name='Arial', size=10)
    data_align = Alignment(vertical='top', wrap_text=True)
    thin_border = Border(bottom=Side(style='thin', color='E0E0E0'))

    def style_header(ws, fill=None):
        f = fill or hdr_fill
        for cell in ws[1]:
            cell.font = hdr_font
            cell.fill = f
            cell.alignment = hdr_align

    def auto_width(ws, max_w=50, min_w=12):
        for col_cells in ws.columns:
            letter = get_column_letter(col_cells[0].column)
            best = max((len(str(c.value).split('\n')[0]) for c in col_cells if c.value), default=min_w)
            ws.column_dimensions[letter].width = max(min_w, min(best + 3, max_w))

    def style_data(ws):
        for r in range(2, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                cell.font = data_font
                cell.alignment = data_align
                cell.border = thin_border

    def clean(v):
        if v is None:
            return ''
        s = str(v).strip()
        return '' if s.lower() == 'nan' else s

    # ── Summary sheet ───────────────────────
    ws = wb.active
    ws.title = 'Summary'
    scols = ['_input', '_status'] + main_fields
    ws.append([HEADER_MAP.get(c, c.replace('_', ' ').title()) for c in scols])
    style_header(ws)

    for row in results:
        ws.append([clean(row.get(c, '')) for c in scols])

    si = scols.index('_status') + 1
    for r in range(2, ws.max_row + 1):
        sc = ws.cell(row=r, column=si)
        if sc.value == 'found':
            sc.value = '✓ Found'
            sc.font = found_font
        elif sc.value == 'not_found':
            sc.value = '✗ Not Found'
            sc.font = notfound_font

    style_data(ws)
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    auto_width(ws)

    # ── Detail sheet builder ────────────────
    def build_sheet(title, detail_key, cols, col_headers, fill=None, hyperlink_col=None, hyperlink_tpl=None, hyperlink_prefix=None):
        if detail_key not in detail_fields:
            return None
        rows_out = []
        for row in results:
            if row.get('_status') != 'found':
                continue
            items = row.get(detail_key, [])
            if not items or not isinstance(items, list):
                continue
            inp = clean(row.get('_input', ''))
            for item in items:
                if not isinstance(item, dict):
                    continue
                vals = [clean(item.get(k, '')) for k in cols]
                if any(vals):
                    rows_out.append([inp] + vals)
        if not rows_out:
            return None

        sheet = wb.create_sheet(title=title)
        sheet.append(['Search Input'] + col_headers)
        style_header(sheet, fill or PatternFill('solid', fgColor='1565C0'))
        for rd in rows_out:
            sheet.append(rd)
        style_data(sheet)

        # Hyperlinks
        if hyperlink_col is not None and hyperlink_tpl:
            ci = hyperlink_col + 2  # +1 for Search Input col, +1 for 1-index
            for r in range(2, sheet.max_row + 1):
                cell = sheet.cell(row=r, column=ci)
                v = str(cell.value or '').strip()
                if v and (not hyperlink_prefix or v.startswith(hyperlink_prefix)):
                    cell.hyperlink = hyperlink_tpl.format(v)
                    cell.font = link_font

        sheet.freeze_panes = 'A2'
        sheet.auto_filter.ref = sheet.dimensions
        auto_width(sheet)
        return sheet

    # ── Genes ───────────────────────────────
    build_sheet('Genes', 'genes_detail',
                ['name', 'id', 'chromosome', 'function'],
                ['Gene Name', 'Gene ID', 'Chromosome', 'Function'])

    # ── Variants ────────────────────────────
    build_sheet('Variants', 'variants_detail',
                ['id', 'type', 'allele', 'gene'],
                ['RS ID', 'Variant Type', 'Allele', 'Gene'],
                hyperlink_col=0,
                hyperlink_tpl='https://www.ncbi.nlm.nih.gov/snp/{}',
                hyperlink_prefix='rs')

    # ── Drugs ───────────────────────────────
    build_sheet('Drugs', 'drugs_detail',
                ['name', 'drug_bank_id', 'state', 'indication'],
                ['Drug Name', 'DrugBank ID', 'State', 'Indication'],
                fill=PatternFill('solid', fgColor='6A1B9A'),
                hyperlink_col=1,
                hyperlink_tpl='https://go.drugbank.com/drugs/{}',
                hyperlink_prefix='DB')

    # ── Phenotypes ──────────────────────────
    build_sheet('Phenotypes', 'phenotypes_detail',
                ['name'],
                ['Phenotype Name'],
                fill=PatternFill('solid', fgColor='E65100'))

    # ── Studies ─────────────────────────────
    build_sheet('Studies', 'studies_detail',
                ['title', 'pubmed_id', 'year', 'study_type'],
                ['Title', 'PubMed ID', 'Year', 'Study Type'],
                fill=PatternFill('solid', fgColor='00695C'),
                hyperlink_col=1,
                hyperlink_tpl='https://pubmed.ncbi.nlm.nih.gov/{}/')

    # ── Countries ───────────────────────────
    build_sheet('Countries', 'countries_detail',
                ['country', 'lat', 'lng'],
                ['Country', 'Latitude', 'Longitude'],
                fill=PatternFill('solid', fgColor='37474F'))

    # ── Export Info sheet ───────────────────
    ws_meta = wb.create_sheet(title='Export Info')
    found_n = sum(1 for r in results if r.get('_status') == 'found')
    meta = [
        ['AGMP Batch Query Export'], [],
        ['Query Type', query_type.title()],
        ['Export Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ['Total Queries', len(results)],
        ['Found', found_n],
        ['Not Found', len(results) - found_n],
        [], ['Sheets Included'],
    ]
    for sn in wb.sheetnames:
        if sn != 'Export Info':
            meta.append(['', sn])
    meta.extend([[], ['Notes'],
        ['', 'Each detail sheet is linked to Summary by the "Search Input" column.'],
        ['', 'Use Excel filters or VLOOKUP to cross-reference across sheets.'],
        ['', 'Hyperlinks to PubMed, dbSNP, and DrugBank are clickable.'],
    ])
    for mr in meta:
        ws_meta.append(mr)
    ws_meta['A1'].font = Font(name='Arial', bold=True, size=14, color='2E7D32')
    for r in range(3, ws_meta.max_row + 1):
        ws_meta.cell(row=r, column=1).font = Font(name='Arial', bold=True, size=10)
        ws_meta.cell(row=r, column=2).font = Font(name='Arial', size=10)
    ws_meta.column_dimensions['A'].width = 20
    ws_meta.column_dimensions['B'].width = 65

    # ── Write and return ────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    ts = datetime.now().strftime('%Y-%m-%d')
    fname = f"batchquery_{query_type}_{ts}.xlsx"
    response = HttpResponse(buf.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    return response

# ============================================================ # ============================================================

# Configure logger
logger = logging.getLogger(__name__)

# Helper function to validate rsID format
def is_valid_rs_id(rs_id):
    """Validate the format of an rsID (typically rs followed by digits)"""
    if not rs_id:
        return False
    rs_pattern = re.compile(r'^rs\d+$')
    return bool(rs_pattern.match(rs_id))

#heatmap-colors
COLORS = {
    "brick_red": "#8B4513",
    "darker_red": "#B22222",
    "lighter_red": "#CD5C5C",
    "orange_yellow": "#FFB266",
    "light_yellow": "#FFFF99",
} 

#current search view
def search_view(request):
    form = ModelSearchForm(request.GET)
    model_selection = ""
    suggestion = None

    if form.is_valid():
        model_selection = form.cleaned_data['model_selection']
        search_query = form.cleaned_data['search_query']

        if model_selection == 'variantagmp':
            results = Variantagmp.objects.filter(rs_id__icontains=search_query).values('rs_id', 'geneagmp__gene_id', 'geneagmp__chromosome').distinct()
            if not results:
                all_variants = Variantagmp.objects.values_list('rs_id', flat=True)
                suggestion = process.extractOne(search_query, all_variants)

        elif model_selection == 'geneagmp':
            results = Geneagmp.objects.filter(gene_id__icontains=search_query).values('gene_id', 'chromosome').distinct()
            if not results:
                all_genes = Geneagmp.objects.values_list('gene_id', flat=True)
                suggestion = process.extractOne(search_query, all_genes)

        elif model_selection == 'drugagmp':
            results = Drugagmp.objects.filter(drug_name__icontains=search_query).values('drug_name', 'drug_id', 'drug_bank_id', 'state', 'indication', 'iupac_name_seq').distinct()
            if not results:
                all_drugs = Drugagmp.objects.values_list('drug_name', flat=True)
                suggestion = process.extractOne(search_query, all_drugs)

        elif model_selection == 'disease':
            results = Variantagmp.objects.select_related().exclude(source_db="PharmGKB").filter(phenotypeagmp__name__icontains=search_query).values("phenotypeagmp__name").distinct()
            if not results:
                all_diseases = Variantagmp.objects.exclude(source_db="PharmGKB").values_list('phenotypeagmp__name', flat=True).distinct()
                suggestion = process.extractOne(search_query, all_diseases)
    else:
        results = []

    return render(request, 'search_list_template.html', {
        'form': form,
        'results': results,
        'model_selection': model_selection,
        'suggestion': suggestion
    })

def search_all(request):
    if request.method == 'POST':
        form = SearchForm(request.POST)
        if form.is_valid():
            search_option = form.cleaned_data['search_option']
            search_query = form.cleaned_data['search_query']
            
            if search_option == 'Variantagmp':
                results = Variantagmp.objects.filter(rs_id__icontains=search_query).values("rs_id","geneagmp__gene_id","geneagmp__chromosome","variant_type").distinct()
            elif search_option == 'Geneagmp':
                results = Geneagmp.objects.filter(gene_id__icontains=search_query)
            elif search_option == 'Drugagmp':
                results = Drugagmp.objects.filter(drug_name__contains=search_query).order_by('drug_name').distinct(F('drug_name').desc())
            elif search_option == 'Disease':
                results = Variantagmp.objects.select_related().exclude(source_db="PharmGKB").filter(phenotypeagmp__name__icontains=search_query).values("phenotypeagmp__name").distinct()
                     
            return render(request, 'search_form.html', {'form': form, 'results': results, 'search_option':search_option})
    else:
        form = SearchForm()
        
    return render(request, 'search_form.html', {'form': form})


#################### Variant Drug Details 1 ################################

class DrugagmpDetailView(DetailView):
    model = Drugagmp
    template_name = 'drugagmp_detail.html'  # Template to display the post details

class VariantStudyagmpListView(ListView):
    model = VariantStudyagmp
    template_name = 'variantstudyagmp_list.html'  # Template to display the comment list

    def get_queryset(self):
        drug_id = self.kwargs['pk']  # Get the post id from URL parameter
        return VariantStudyagmp.objects.filter(Variantagmp__drugagmp_icontains=drug_id)  # Filter comments by post id


  
#################### Variant Drug Details ################################
  
#################### PharmacoGene Associations exclude Gwas catalogue in the first queryset ################################
#03 Drug associations and Phenotype Associations
class PhamacogeneDrugAssoc(DetailView):
    model = VariantStudyagmp
    template_name = 'PhamacogeneDrugAssoc.html'
    pk_url_kwarg = 'gene_id'
    context_object_name = 'variantstudyagmp'


    def get_object(self):
        gene_id = self.kwargs.get(self.pk_url_kwarg)
        
        try:
            data = Geneagmp.objects.filter(gene_id=gene_id)
            if not data.exists():
                logger.warning(f"No gene found with gene_id: {gene_id}")
                raise Http404(f"No gene found with gene_id: {gene_id}")
            return data
        except Exception as e:
            logger.error(f"Error retrieving gene {gene_id}: {str(e)}")
            raise Http404(f"Error retrieving gene: {str(e)}")
    
    def get_context_data(self, **kwargs):
        context = super(PhamacogeneDrugAssoc, self).get_context_data(**kwargs)
        gene_id = self.kwargs.get(self.pk_url_kwarg)

        try:
            #content to display
            context['geneagmp'] = Geneagmp.objects.filter(
                gene_id=gene_id)
            
            context["data"] = Geneagmp.objects.filter(gene_id = gene_id).first()
           
            #include PharmGKB and Exclude Gwas Catalogue & DisGeNET
            context['object_list'] = VariantStudyagmp.objects.filter(
                variantagmp__geneagmp__gene_id__iregex=r"(^|[^a-zA-Z0-9_]){0}([^a-zA-Z0-9_]|$)".format(
                    re.escape(str(gene_id)).replace('\\ ', '\\s+')
                )
            ).exclude(
                variantagmp__source_db__in=["DisGeNET", "GWAS Catalog"])
            
            context['object_list_diseases'] = VariantStudyagmp.objects.filter(
                variantagmp__geneagmp__gene_id__iregex=r"(^|[^a-zA-Z0-9_]){0}([^a-zA-Z0-9_]|$)".format(re.escape(str(gene_id))),
                variantagmp__source_db__in=["DisGeNet", "GWAS Catalog"]
            ).exclude(
                variantagmp__source_db="PharmGKB"
            )
        except Exception as e:
            logger.error(f"Error in get_context_data for gene {gene_id}: {str(e)}")
            context['error'] = "An error occurred while retrieving data."
        
        return context

#################### Gene Drug Associations ################################


#################### Var Drug Associations exclude gwas catalogue ################################
class VarDrugAssocDetailView(DetailView):
    model = VariantStudyagmp
    template_name = 'VarDrugAssocDetail.html'
    pk_url_kwarg = 'rs_id'
    context_object_name = 'variantstudyagmp'

    def get_object(self):
        rs_id = self.kwargs.get(self.pk_url_kwarg)
        
        # Validate rs_id format first
        if rs_id and not rs_id.startswith("DB") and not is_valid_rs_id(rs_id):
            logger.warning(f"Invalid rs_id format: {rs_id}")
            raise Http404(f"Invalid variant ID format: {rs_id}")

        try:
            data = Variantagmp.objects.filter(rs_id=rs_id)
            if not data.exists():
                logger.warning(f"No variant found with rs_id: {rs_id}")
                raise Http404(f"No variant found with rs_id: {rs_id}")
            return data
        except Exception as e:
            logger.error(f"Error retrieving variant {rs_id}: {str(e)}")
            raise Http404(f"Error retrieving variant: {str(e)}")
    
    def get_context_data(self, **kwargs):
        context = super(VarDrugAssocDetailView, self).get_context_data(**kwargs)
        rs_id = self.kwargs.get(self.pk_url_kwarg)
        
        try:  
            context['gene_id_display'] = Variantagmp.objects.values("geneagmp__gene_id").filter(rs_id=rs_id).first()
            context['chromosome_display'] = Variantagmp.objects.values("geneagmp__chromosome").filter(rs_id=rs_id).first()
            context['rs_id_display'] = Variantagmp.objects.values("rs_id").filter(rs_id=rs_id).first()
      
            #back up query
            context['object_list'] = VariantStudyagmp.objects.filter(
                variantagmp__rs_id__iregex=r"\b{0}\b".format(str(rs_id))).exclude(variantagmp__source_db="DisGeNET").exclude(variantagmp__source_db="GWAS Catalog")
        except Exception as e:
            logger.error(f"Error in get_context_data for variant {rs_id}: {str(e)}")
            context['error'] = "An error occurred while retrieving data."
        
        return context

#################### Variant Disease Associations ################################
# 02 
class VariantDiseaseAssocDetailView(DetailView):
    model = VariantStudyagmp
    template_name = 'VariantDiseaseAssocDetail.html'
    pk_url_kwarg = 'rs_id'

    def get_object(self):
        rs_id = self.kwargs.get(self.pk_url_kwarg)
        
        # Validate rs_id format first
        if rs_id and not rs_id.startswith("DB") and not is_valid_rs_id(rs_id):
            logger.warning(f"Invalid rs_id format: {rs_id}")
            raise Http404(f"Invalid variant ID format: {rs_id}")

        try:
            data = Variantagmp.objects.filter(rs_id=rs_id)
            if not data.exists():
                logger.warning(f"No variant found with rs_id: {rs_id}")
                raise Http404(f"No variant found with rs_id: {rs_id}")
            return data
        except Exception as e:
            logger.error(f"Error retrieving variant {rs_id}: {str(e)}")
            raise Http404(f"Error retrieving variant: {str(e)}")
    
    def get_context_data(self, **kwargs):
        context = super(VariantDiseaseAssocDetailView, self).get_context_data(**kwargs)
        rs_id = self.kwargs.get(self.pk_url_kwarg)
     
        try:
            if Variantagmp.objects.filter(rs_id=rs_id).exists():
                context['rs_id_display'] = (Variantagmp.objects.values("rs_id").filter(rs_id=rs_id))[0]
                context['gene_name_display'] = Variantagmp.objects.values("geneagmp__gene_id").filter(rs_id=rs_id).first()
                context['chromosome_display'] = Variantagmp.objects.values("geneagmp__chromosome").filter(rs_id=rs_id).first()

                context['object_list'] = VariantStudyagmp.objects.filter(
                    variantagmp__rs_id__iregex=r"(^|[^a-zA-Z0-9_]){0}([^a-zA-Z0-9_]|$)".format(re.escape(str(rs_id))),
                    variantagmp__source_db__iregex=r"^(DisGeNet|GWAS Catalog)$"
                ).exclude(variantagmp__source_db="PharmGKB")
            else:
                context['error'] = f"No information found for variant: {rs_id}"
        except Exception as e:
            logger.error(f"Error in get_context_data for variant {rs_id}: {str(e)}")
            context['error'] = "An error occurred while retrieving data."

        return context
    

#################### DRUG searchs for Variant drug Associations ################################
#01
class VariantDrugAssociationDetailView(DetailView):
    model = Variantagmp
    pk_url_kwarg = 'rs_id'
    
    def get_template_names(self):
        rs_id = self.kwargs.get(self.pk_url_kwarg)
        if rs_id and rs_id.startswith("DB"):
            return ['VariantDrugAssociation.html']
        else:
            return ['VarDrugAssocDetail.html']
    
    def get_object(self):
        rs_id = self.kwargs.get(self.pk_url_kwarg)
        
        # If not a DB ID, validate the rs_id format first
        if rs_id and not rs_id.startswith("DB") and not is_valid_rs_id(rs_id):
            logger.warning(f"Invalid rs_id format: {rs_id}")
            raise Http404(f"Invalid variant ID format: {rs_id}")
            
        # If the ID starts with "DB", look up by drug ID instead
        if rs_id and rs_id.startswith("DB"):
            try:
                # Find a variant associated with this drug
                drug = get_object_or_404(Drugagmp, drug_bank_id=rs_id)
                variant = Variantagmp.objects.filter(drugagmp=drug).first()
                if variant:
                    return variant
                else:
                    # Still return something even if no variant is found
                    # This allows the view to continue and show drug info
                    return Variantagmp(drugagmp=drug)
            except Http404:
                logger.warning(f"No drug found with drug_bank_id: {rs_id}")
                raise Http404(f"No drug found with ID: {rs_id}")
            except Exception as e:
                logger.error(f"Error retrieving drug {rs_id}: {str(e)}")
                raise Http404(f"Error retrieving drug: {str(e)}")
        else:
            # For non-DB IDs - Always get the first matching variant when multiple exist
            try:
                # First try with a direct filter
                variants = Variantagmp.objects.filter(rs_id=rs_id)
                if variants.exists():
                    return variants.first()  # Return the first match if found
                else:
                    # If no variants found, raise a 404 instead of causing a 500 error
                    logger.warning(f"No variant found with rs_id: {rs_id}")
                    raise Http404(f"No variant found with rs_id: {rs_id}")
            except Variantagmp.DoesNotExist:
                # Also handle the DoesNotExist exception properly
                logger.warning(f"Variant.DoesNotExist error for rs_id: {rs_id}")
                raise Http404(f"No variant found with rs_id: {rs_id}")
            except Exception as e:
                # Catch any other unexpected errors and log them
                logger.error(f"Unexpected error for rs_id {rs_id}: {str(e)}")
                raise Http404(f"Error retrieving variant: {str(e)}")
                
    def get_context_data(self, **kwargs):
        context = super(VariantDrugAssociationDetailView, self).get_context_data(**kwargs)
        rs_id = self.kwargs.get(self.pk_url_kwarg)
        
        try:
            # If this is a drug ID, add drug data directly to context
            if rs_id and rs_id.startswith("DB"):
                drug = get_object_or_404(Drugagmp, drug_bank_id=rs_id)
                context['data'] = drug
                # Get all variant studies related to this drug
                variant_studies = VariantStudyagmp.objects.filter(
                    variantagmp__drugagmp=drug
                ).select_related(
                    'variantagmp',
                    'studyagmp',
                    'variantagmp__drugagmp',
                    'variantagmp__geneagmp'
                )
                context['object_list'] = variant_studies
                # If we need an rs_id for other parts of the template, get it from the first variant
                variant = Variantagmp.objects.filter(drugagmp=drug).first()
                if variant:
                    rs_id = variant.rs_id
                else:
                    rs_id = None
            else:
                # If it's a variant ID, get associated drug info
                variant = self.object
                if variant and hasattr(variant, 'drugagmp') and variant.drugagmp:
                    context['data'] = variant.drugagmp
                    # Get all variant studies for this variant
                    context['object_list'] = VariantStudyagmp.objects.filter(
                        variantagmp__rs_id=variant.rs_id
                    ).select_related(
                        'variantagmp',
                        'studyagmp',
                        'variantagmp__drugagmp',
                        'variantagmp__geneagmp'
                    ).exclude(
                        variantagmp__source_db__in=["DisGeNET", "GWAS Catalog"]
                    )
            
            # Get the variant object for display in the template
            variant = self.object
            # Add variant info to context
            context['rs_id_display'] = variant
            # Get gene information for this variant
            if variant and hasattr(variant, 'geneagmp') and variant.geneagmp:
                context['gene_id_display'] = {
                    'geneagmp__gene_id': variant.geneagmp.gene_id
                }
                context['chromosome_display'] = {
                    'geneagmp__chromosome': variant.geneagmp.chromosome
                }
        except Exception as e:
            logger.error(f"Error in context data for {rs_id}: {str(e)}")
            context['error'] = "An error occurred while retrieving data."
            
        return context

#################### Variant Var Drug Associations ################################
class VvarDrugAssocDetailView(DetailView):
    model = VariantStudyagmp
    template_name = 'VarDrugAssocDetail.html'
    pk_url_kwarg = 'rs_id'

    def get_object(self):
        rs_id = self.kwargs.get(self.pk_url_kwarg)
        
        # Validate rs_id format first
        if rs_id and not rs_id.startswith("DB") and not is_valid_rs_id(rs_id):
            logger.warning(f"Invalid rs_id format: {rs_id}")
            raise Http404(f"Invalid variant ID format: {rs_id}")
            
        try:
            data = Variantagmp.objects.filter(rs_id=rs_id)
            if not data.exists():
                logger.warning(f"No variant found with rs_id: {rs_id}")
                raise Http404(f"No variant found with rs_id: {rs_id}")
            return data
        except Exception as e:
            logger.error(f"Error retrieving variant {rs_id}: {str(e)}")
            raise Http404(f"Error retrieving variant: {str(e)}")
    
    def get_context_data(self, **kwargs):
        context = super(VvarDrugAssocDetailView, self).get_context_data(**kwargs)
        rs_id = self.kwargs.get(self.pk_url_kwarg)
        
        try:
            context['variantagmp'] = Variantagmp.objects.filter(rs_id=rs_id)
            variant = Variantagmp.objects.filter(rs_id=rs_id)
            
            context['object_list'] = VariantStudyagmp.objects.filter(
                variantagmp__rs_id__iregex=r"\b{0}\b".format(str(rs_id)))
        except Exception as e:
            logger.error(f"Error in get_context_data for variant {rs_id}: {str(e)}")
            context['error'] = "An error occurred while retrieving data."
            
        return context

#################### Search Diseases ################################

# Display Phamacogenes and Disease associations
#04
class DiseaseVariantDetailView(DetailView):
    model = VariantStudyagmp
    template_name = 'DiseaseVariantDetailView.html'
    pk_url_kwarg = 'phenotypeagmp__name'


    def get_object(self):
        phenotypeagmp__name = self.kwargs.get(self.pk_url_kwarg)
        
        try:
            # Check if data exists for this phenotype
            variants = Variantagmp.objects.filter(phenotypeagmp__name=phenotypeagmp__name)
            if not variants.exists():
                logger.warning(f"No variants found with phenotype: {phenotypeagmp__name}")
                raise Http404(f"No variants found with phenotype: {phenotypeagmp__name}")
            # This method doesn't actually return anything, but we need to pass the check
            return variants
        except Exception as e:
            logger.error(f"Error checking phenotype {phenotypeagmp__name}: {str(e)}")
            raise Http404(f"Error retrieving phenotype data: {str(e)}")

    def get_context_data(self, **kwargs):
        context = super(DiseaseVariantDetailView, self).get_context_data(**kwargs)
        phenotypeagmp__name = self.kwargs.get(self.pk_url_kwarg)

        try:
            phenotype_data = Variantagmp.objects.filter(
                phenotypeagmp__name=phenotypeagmp__name).values("phenotypeagmp__name").distinct()
                
            if phenotype_data.exists():
                context['data'] = phenotype_data[0]
               
                context['object_list1'] = VariantStudyagmp.objects.select_related().filter(
                    variantagmp__phenotypeagmp__name__iregex=r"\\y{0}\\y".format(str(phenotypeagmp__name))
                ).exclude(variantagmp__source_db="PharmGKB")
               
                context['object_list'] = VariantStudyagmp.objects.select_related().filter(
                    variantagmp__phenotypeagmp__name__iexact=phenotypeagmp__name).exclude(variantagmp__source_db="PharmGKB")
            else:
                context['error'] = f"No data found for phenotype: {phenotypeagmp__name}"
        except Exception as e:
            logger.error(f"Error in get_context_data for phenotype {phenotypeagmp__name}: {str(e)}")
            context['error'] = "An error occurred while retrieving data."
        
        return context
       
#################### Variant Var Drug Associations ################################
class VarDisAssocDetailView(DetailView):
    model = VariantStudyagmp
    template_name = 'VarDissAssocDetail.html'
    pk_url_kwarg = 'rs_id'

    def get_object(self):
        rs_id = self.kwargs.get(self.pk_url_kwarg)
        
        # Validate rs_id format first
        if rs_id and not rs_id.startswith("DB") and not is_valid_rs_id(rs_id):
            logger.warning(f"Invalid rs_id format: {rs_id}")
            raise Http404(f"Invalid variant ID format: {rs_id}")
            
        try:
            data = Variantagmp.objects.filter(rs_id=rs_id)
            if not data.exists():
                logger.warning(f"No variant found with rs_id: {rs_id}")
                raise Http404(f"No variant found with rs_id: {rs_id}")
            return data
        except Exception as e:
            logger.error(f"Error retrieving variant {rs_id}: {str(e)}")
            raise Http404(f"Error retrieving variant: {str(e)}")
    
    def get_context_data(self, **kwargs):
        context = super(VarDisAssocDetailView, self).get_context_data(**kwargs)
        rs_id = self.kwargs.get(self.pk_url_kwarg)
     
        try:
            context['variantagmp'] = Variantagmp.objects.filter(rs_id=rs_id)
            variant = Variantagmp.objects.filter(rs_id=rs_id)
            
            context['object_list'] = VariantStudyagmp.objects.filter(
                variantagmp__rs_id__iregex=r"\b{0}\b".format(str(rs_id))).exclude(variantagmp__source_db="PharmGKB")
        except Exception as e:
            logger.error(f"Error in get_context_data for variant {rs_id}: {str(e)}")
            context['error'] = "An error occurred while retrieving data."
        
        return context

# Display Phamacogenes and Disease associations
class PharmacoDrugDetailView(DetailView):
    model = VariantStudyagmp
    template_name = 'PharmacoDrugDetailView.html'
    pk_url_kwarg = 'gene_id'
    

    def get_object(self):
        gene_id = self.kwargs.get(self.pk_url_kwarg)

        try:
            data = Geneagmp.objects.filter(gene_id=gene_id)
            if not data.exists():
                logger.warning(f"No gene found with gene_id: {gene_id}")
                raise Http404(f"No gene found with gene_id: {gene_id}")
            return data
        except Exception as e:
            logger.error(f"Error retrieving gene {gene_id}: {str(e)}")
            raise Http404(f"Error retrieving gene: {str(e)}")
    
    def get_context_data(self, **kwargs):
        context = super(PharmacoDrugDetailView, self).get_context_data(**kwargs)
        gene_id = self.kwargs.get(self.pk_url_kwarg)

        try:
            context['geneagmp'] = Geneagmp.objects.filter(
                gene_id=gene_id).first()
            
            context['object_list'] = VariantStudyagmp.objects.filter(
                variantagmp__geneagmp__gene_id__iregex=r"\b{0}\b".format(str(gene_id))) 
            
            context['object_list_diseases_old']=VariantStudyagmp.objects.select_related().filter(variantagmp__geneagmp__gene_id__icontains=gene_id)

            context['object_list_diseases'] = VariantStudyagmp.objects.select_related().filter(variantagmp__geneagmp__gene_id__iregex=r"\b{0}\b".format(str(gene_id))).exclude(variantagmp__source_db="PharmGKB")
        except Exception as e:
            logger.error(f"Error in get_context_data for gene {gene_id}: {str(e)}")
            context['error'] = "An error occurred while retrieving data."
        
        return context


#################### Variant Drug Details ################################

class DrugDetailView(DetailView):
    model = VariantStudyagmp
    template_name = 'drug_detail.html'
    pk_url_kwarg = 'drug_id'

    def get_object(self):
        drug_id = self.kwargs.get(self.pk_url_kwarg)

        try:
            data = Drugagmp.objects.filter(drug_id=drug_id)
            if not data.exists():
                logger.warning(f"No drug found with drug_id: {drug_id}")
                raise Http404(f"No drug found with drug_id: {drug_id}")
            return data
        except Exception as e:
            logger.error(f"Error retrieving drug {drug_id}: {str(e)}")
            raise Http404(f"Error retrieving drug: {str(e)}")
    
    def get_context_data(self, **kwargs):
        context = super(DrugDetailView, self).get_context_data(**kwargs)
        drug_id = self.kwargs.get(self.pk_url_kwarg)
        
        try:
            context['drugagmp'] = Drugagmp.objects.filter(
                drug_id=drug_id).first()
            drug = Drugagmp.objects.filter(drug_id=drug_id).first()

            if drug:
                context['object_list'] = VariantStudyagmp.objects.filter(
                    variantagmp__drugagmp__drug_id__iregex=r"\b{0}\b".format(str(drug_id)))
                
                context['drugagmp'] = Drugagmp.objects.filter(
                   drug_id=drug.id).first()
            else:
                context['error'] = f"No drug found with ID: {drug_id}"
        except Exception as e:
            logger.error(f"Error in get_context_data for drug {drug_id}: {str(e)}")
            context['error'] = "An error occurred while retrieving data."
            
        return context

def about(request):
    return render(request, 'about.html')

def get_map_data(request, map_type):
    """
    AJAX endpoint for getting map data based on study type filter.
    - marker:  returns {success, points: [{lat, lng, count}]}
    - heatmap: returns {success, countries: [{country, count}]}
               (point-in-polygon aggregation so the client can shade full borders)
    """
    study_type = request.GET.get('study_type', 'All')

    try:
        def get_filtered_studies(st):
            studies = VariantStudyagmp.objects.select_related('studyagmp').distinct('studyagmp__publication_id')
            if st and st != 'All':
                studies = studies.filter(studyagmp__study_type=st)
            return studies

        def get_location_data(lat_field, lon_field, queryset):
            return queryset.exclude(
                Q(**{f'{lon_field}__isnull': True}) | Q(**{f'{lon_field}__exact': ''}) |
                Q(**{f'{lat_field}__isnull': True}) | Q(**{f'{lat_field}__exact': ''}) |
                Q(**{f'{lon_field}__iexact': 'nan'}) | Q(**{f'{lat_field}__iexact': 'nan'})
            ).values('studyagmp__publication_id', lat_field, lon_field).annotate(
                _lat=F(lat_field),
                _lng=F(lon_field)
            ).values('_lat', '_lng')

        # Model has an unsuffixed base pair plus _01 through _30
        location_fields = [('latitude', 'longitude')] + [
            (f'latitude_{i:02d}', f'longitude_{i:02d}') for i in range(1, 31)
        ]

        filtered_studies = get_filtered_studies(study_type)
        locations = [get_location_data(lat, lon, filtered_studies) for lat, lon in location_fields]
        flattened_locations = [item for sublist in locations for item in sublist]

        # Deduplicate by coordinate
        count_per_coordinates = defaultdict(int)
        for record in flattened_locations:
            coordinates = (record["_lat"], record["_lng"])
            count_per_coordinates[coordinates] += 1

        # Clean: skip NaN / Infinity values (they break JSON serialisation)
        clean_points = []
        for coordinates, value in count_per_coordinates.items():
            try:
                lat = float(coordinates[0])
                lng = float(coordinates[1])
                if math.isnan(lat) or math.isnan(lng) or math.isinf(lat) or math.isinf(lng):
                    continue
                clean_points.append((lat, lng, value))
            except (ValueError, TypeError):
                continue

        if map_type == 'marker':
            points = [{'lat': lat, 'lng': lng, 'count': cnt} for lat, lng, cnt in clean_points]
            return JsonResponse({'success': True, 'points': points})

        elif map_type == 'heatmap':
            # Aggregate points into countries using the local GeoJSON boundaries
            geojson_path = os.path.join(settings.BASE_DIR, 'agmp_app/static/maps/countries.geo.json')
            gdf = gpd.read_file(geojson_path)

            publications_per_country = defaultdict(int)
            for lat, lng, value in clean_points:
                try:
                    point = Point(float(lng), float(lat))
                    for _, row in gdf.iterrows():
                        if row['geometry'].contains(point):
                            publications_per_country[row['name']] += value
                            break
                except (ValueError, TypeError):
                    continue

            countries = [
                {'country': name, 'count': count}
                for name, count in sorted(publications_per_country.items(), key=lambda x: -x[1])
            ]
            return JsonResponse({'success': True, 'countries': countries})

        return JsonResponse({'success': False, 'error': f'Unknown map_type: {map_type}'}, status=400)

    except Exception as e:
        logger.error(f"Error generating map data: {str(e)}")
        return JsonResponse({'success': False, 'error': 'An error occurred while generating map data'}, status=500)
    

def summary(request):
    """
    Main view for the summary page
    """
    # Basic counts
    unique_genes = Geneagmp.objects.exclude(gene_id__iexact='').exclude(gene_id__iexact="nan").values('gene_id').distinct()
    gene_count = unique_genes.count()
    drug_count = Drugagmp.objects.exclude(drug_bank_id__iexact='').exclude(drug_bank_id__iexact="nan").values('drug_bank_id').distinct().count()
    variant_count = Variantagmp.objects.exclude(rs_id__iexact='').exclude(rs_id__iexact="nan").values('rs_id').distinct().count()
    disease_count = Variantagmp.objects.values('phenotypeagmp__name').distinct().count()
    publication_count = Studyagmp.objects.exclude(publication_id__iexact='').exclude(publication_id__iexact="nan").values('publication_id').distinct().count()
    
    # Optimized queries for graphs
    qs_drug = (
        Drugagmp.objects.exclude(drug_name="nan")
        .values('drug_name')
        .annotate(frequency=Count('drugs'))
        .order_by('-frequency')[:10]
    )
    
    qs_gene = (
        Geneagmp.objects.exclude(gene_name="nan")
        .values('gene_id')
        .annotate(frequency=Count('variantagmp__studyagmp'))
        .order_by('-frequency')[:10]
    )
    
    qs_variant = (
        Variantagmp.objects.exclude(rs_id="nan")
        .values('rs_id')
        .annotate(frequency=Count('studyagmp'))
        .order_by('-frequency')[:10]
    )
    
    qs_disease = (
        Phenotypeagmp.objects.exclude(variantagmp__source_db="PharmGKB")
        .exclude(variantagmp__source_db="nan")
        .values('name')
        .annotate(frequency=Count('variantagmp'))
        .order_by('-frequency')[:10]
    )
    
    context = {
        'gene_count': gene_count,
        'publication_count': publication_count,
        'drug_count': drug_count,
        'variant_count': variant_count,
        'disease_count': disease_count,
        'qs_drug': qs_drug,
        'qs_gene': qs_gene,
        'qs_variant': qs_variant,
        'qs_disease': qs_disease,
        'study_types': ['All', 'GWAS', 'Case Report','Candidate Gene','WES/WGS','Clinical Trial','Other']
    }
    
    return render(request, 'summary.html', context)



def outreach(request):
    return render(request, 'outreach.html')

def contact(request):
    return render(request, 'contact.html')

def disclaimer(request):
    return render(request, 'disclaimer.html')

def faqs(request):
    return render(request, 'faqs.html')

def tools_pipelines(request):
    return render(request, 'tools_pipelines.html')

def databases(request):
    return render(request, 'resources.html')

def online_courses(request):
    return render(request, 'online_courses.html')

def help(request):
    return render(request, 'help.html')

def tutorial(request):
    return render(request, 'tutorial.html')

def home(request):
    return render(request, 'home.html')

def test_data_table(request):
    return render(request, 'test_data_table.html')

from django.http import JsonResponse
from collections import defaultdict
import reverse_geocoder as rg

# Local cache for lat/lon lookups
coord_cache = {}

def is_valid_coord(lat, lon):
    """
    Checks whether lat/lon are valid values (not blank, not NaN).
    """
    try:
        return (
            lat is not None
            and lon is not None
            and lat != ""
            and lon != ""
            and str(lat).lower() != "nan"
            and str(lon).lower() != "nan"
        )
    except Exception:
        return False

def reverse_geocode(lat, lon):
    """
    Reverse-geocode lat/lon to country code.
    Caches results for efficiency.
    """
    key = (round(float(lat), 4), round(float(lon), 4))
    if key in coord_cache:
        return coord_cache[key]
    
    result = rg.search(key, mode=1)[0]
    country_code = result["cc"]
    coord_cache[key] = country_code
    return country_code

def studies_per_country(request):
    """
    Returns JSON of unique publication_ids per country.
    """

    from .models import VariantStudyagmp

    variant_studies = VariantStudyagmp.objects.all()

    # Mapping: country → set of publication_ids
    country_to_publication_ids = defaultdict(set)

    for vs in variant_studies:
        for i in range(12):
            country_field = "country_participant" if i == 0 else f"country_participant_{i:02}"
            lat_field = "latitude" if i == 0 else f"latitude_{i:02}"
            lon_field = "longitude" if i == 0 else f"longitude_{i:02}"

            country_name = getattr(vs, country_field, None)
            lat = getattr(vs, lat_field, None)
            lon = getattr(vs, lon_field, None)

            # Determine country
            if country_name:
                country = country_name.strip()
            elif is_valid_coord(lat, lon):
                try:
                    country = reverse_geocode(lat, lon)
                except Exception:
                    continue
            else:
                continue

            # Ensure study exists and has a publication_id
            if (
                vs.studyagmp
                and vs.studyagmp.publication_id
                and vs.studyagmp.publication_id.strip() != ""
            ):
                pub_id = vs.studyagmp.publication_id.strip()
                country_to_publication_ids[country].add(pub_id)

    # Prepare final result: count unique publication_ids per country
    result = {
        country: len(publication_ids)
        for country, publication_ids in country_to_publication_ids.items()
    }

    return JsonResponse(result)

from django.shortcuts import render
from collections import defaultdict
import reverse_geocoder as rg

coord_cache = {}

def is_valid_coord(lat, lon):
    """
    Checks whether lat/lon are valid values (not blank, not NaN).
    """
    try:
        return (
            lat is not None
            and lon is not None
            and lat != ""
            and lon != ""
            and str(lat).lower() != "nan"
            and str(lon).lower() != "nan"
        )
    except Exception:
        return False

def reverse_geocode(lat, lon):
    """
    Reverse-geocode lat/lon to country code.
    Caches results for efficiency.
    """
    key = (round(float(lat), 4), round(float(lon), 4))
    if key in coord_cache:
        return coord_cache[key]
    
    result = rg.search(key, mode=1)[0]
    country_code = result["cc"]
    coord_cache[key] = country_code
    return country_code